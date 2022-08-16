from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl, time, threading, random, string

# Signin => False / Signup => True
signup_flag = False

# Number of Users for test ( auto / enter number )
auto = True
users_no = 5
min_range = 0
max_range = 0

# Sleep Between Process
sleep_between_process = 1

# Sleep Between Threads Process
sleep_between_thread = 0.5

# Browsers Path
chrome_path = "path to chrome driver file"
firefox_path = "path to firefox driver file"

# Headless Browser
headless_browser = False

# List of Thread
signup_thread_list = []
signin_thread_list = []

# Get Ranges
if auto:
    min_range = int(input("Enter users number minimum range : "))
    max_range = int(input("Enter users number maximum range : "))


def signin_test_logic(email, password, r):
    # Browser Configuration
    driver = Service(chrome_path)
    option = webdriver.ChromeOptions()
    option.headless = headless_browser
    browser = webdriver.Chrome(service=driver, options=option)
    url = "YOUR WEBSITE URL"
    browser.get(url)
    assert url in browser.current_url
    print(f"\n ### {r} Home Page Loaded. ###")

    # Signin
    signin_form = browser.find_element(By.CLASS_NAME, "navbar-brand")
    signin_form.click()
    signin_url = "YOUR WEBSITE LOGIN URL"
    assert signin_url in browser.current_url
    print(f" ### {r} Signin Page Loaded. ###")

    # Find Login Elements
    time.sleep(sleep_between_process)
    signup_form = browser.find_element(By.CLASS_NAME, "sign-up-htm")
    time.sleep(sleep_between_process)
    if signup_form.is_displayed():
        username_input = browser.find_element(By.ID, "user_login")
        password_input = browser.find_element(By.ID, "pass_login")
        submit_button = browser.find_element(By.ID, "btn_login")
        # Send Data
        time.sleep(sleep_between_process)
        username_input.send_keys(email)
        password_input.send_keys(password)
        time.sleep(sleep_between_process)
        submit_button.click()

        # Check User Signin Successfully
        home_url = "YOUR WEBSITE HOME URL"
        if home_url == browser.current_url:
            print(" ### User Singed in Successfully. ###")
        else:
            print(" ### Error!! (Signin error) ###")

    # Get Response Time
    """
    Performance Timing Events flow
    navigationStart -> redirectStart -> redirectEnd -> fetchStart -> domainLookupStart -> domainLookupEnd
    -> connectStart -> connectEnd -> requestStart -> responseStart -> responseEnd
    -> domLoading -> domContentLoaded -> domComplete -> loadEventStart -> loadEventEnd
    """
    navigation_start = browser.execute_script("return window.performance.timing.navigationStart")
    response_start = browser.execute_script("return window.performance.timing.responseStart")
    dom_complete = browser.execute_script("return window.performance.timing.domComplete")

    backend_performance = response_start - navigation_start
    frontend_performance = dom_complete - response_start

    print(
        f"| {r}_backend_performance => {backend_performance} ms \n | {r}_frontend_performance => {frontend_performance} ms")
    time.sleep(sleep_between_process)
    browser.close()


def signup_test_logic(email, password, r):
    # Browser Configuration
    driver = Service(chrome_path)
    option = webdriver.ChromeOptions()
    option.headless = headless_browser
    browser = webdriver.Chrome(service=driver, options=option)
    url = "YOUR WEBSITE URL"
    browser.get(url)
    assert url in browser.current_url
    print(f"\n ### {r} Home Page Loaded. ###")

    # SignUp
    signup_form = browser.find_element(By.CLASS_NAME, "navbar-brand")
    signup_form.click()
    signup_url = "YOUR WEBSITE REGISTER URL"
    assert signup_url in browser.current_url
    print(f" ### {r} Signup Page Loaded. ###")

    # Find SignUp Elements
    time.sleep(sleep_between_process)
    browser.find_element(By.XPATH, "/html/body/div/div/label[2]").click()
    signup_form = browser.find_element(By.CLASS_NAME, "sign-up-htm")
    time.sleep(sleep_between_process)
    if signup_form.is_displayed():
        username_input = browser.find_element(By.ID, "user_signup")
        password_input = browser.find_element(By.ID, "pass_sign")
        firstname_input = browser.find_element(By.NAME, "firstname")
        submit_button = browser.find_element(By.XPATH, "/html/body/div/div/div/div[2]/form/div[4]/input")
        # Send Data
        time.sleep(sleep_between_process)
        username_input.send_keys(email)
        password_input.send_keys(password)
        firstname_input.send_keys("".join(random.choice(string.ascii_letters) for i in range(10)))
        time.sleep(sleep_between_process)
        submit_button.click()

    # Check User Signed up or not
    home_url = "YOUR WEBSITE HOME URL"
    if home_url == browser.current_url:
        print(" ### Insert new user in database and redirect to home page ###")
    else:
        print(" ### User Exist!! ###")

    # Get Response Time
    """
    Performance Timing Events flow
    navigationStart -> redirectStart -> redirectEnd -> fetchStart -> domainLookupStart -> domainLookupEnd
    -> connectStart -> connectEnd -> requestStart -> responseStart -> responseEnd
    -> domLoading -> domContentLoaded -> domComplete -> loadEventStart -> loadEventEnd
    """
    navigation_start = browser.execute_script("return window.performance.timing.navigationStart")
    response_start = browser.execute_script("return window.performance.timing.responseStart")
    dom_complete = browser.execute_script("return window.performance.timing.domComplete")

    backend_performance = response_start - navigation_start
    frontend_performance = dom_complete - response_start

    print(f"| {r}_backend_performance => {backend_performance} ms \n | {r}_frontend_performance => {frontend_performance} ms")
    time.sleep(sleep_between_process)
    browser.close()


def get_users_from_excel():
    # Get Data from Excel sheet
    excel_path = "PATH TO EXCEL FILE"
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    return sheet


def signup_threads_run():
    sheet = get_users_from_excel()
    for r in range(2, users_no+2):
        row = random.randint(min_range + 1, max_range)
        email = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        # Call signup_test_logic with Threading
        thread = threading.Thread(target=signup_test_logic, args=(email, password, r-1))
        time.sleep(sleep_between_thread)
        thread.start()
        print(f"Thread {r - 1} is running...")
        signup_thread_list.append(thread)


def signin_threads_run():
    sheet = get_users_from_excel()
    for r in range(2, users_no+2):
        row = random.randint(min_range + 1, max_range)
        email = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        # Call signin_test_logic with Threading
        thread = threading.Thread(target=signin_test_logic, args=(email, password, r-1))
        time.sleep(sleep_between_thread)
        thread.start()
        print(f"Thread {r-1} is running...")
        signin_thread_list.append(thread)


def generate_random_number(users_no_min_range, users_no_max_range):
    global users_no
    users_no = random.randint(users_no_min_range, users_no_max_range)
    print(f"* users_no => {users_no} *")


def run_signup_test():
    if auto:
        generate_random_number(min_range, max_range)
    signup_threads_run()
    for thread in signup_thread_list:
        thread.join()
    print("------------------------ next signup test --------------------------------")
    run_signup_test()


def run_signin_test():
    if auto:
        generate_random_number(min_range, max_range)
    signin_threads_run()
    for thread in signin_thread_list:
        thread.join()
    print("------------------------ next signin test --------------------------------")
    run_signin_test()


if signup_flag:
    run_signup_test()
else:
    run_signin_test()
